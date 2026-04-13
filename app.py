import streamlit as st
import pandas as pd
import numpy as np
import unicodedata
import re
import os
import subprocess
import tempfile
import zipfile
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

st.set_page_config(
    page_title="Anexo ANMAT Natura",
    page_icon="📋",
    layout="wide"
)

# ─────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────
st.markdown("""
<style>
    .header-box {
        background: linear-gradient(135deg, #e8f4f8 0%, #d0eaf5 100%);
        border-left: 5px solid #00b4d8;
        border-radius: 8px;
        padding: 20px 28px;
        margin-bottom: 24px;
    }
    .header-box h1 { color: #00b4d8; font-size: 1.6rem; font-weight: 700; margin: 0 0 4px 0; }
    .header-box p { color: #444; font-size: 0.9rem; margin: 0; }
    .card { background: #ffffff; border: 1px solid #dde3ea; border-radius: 8px; padding: 20px; margin-bottom: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }
    .card h3 { color: #00b4d8; font-size: 0.9rem; font-weight: 700; margin: 0 0 14px 0; text-transform: uppercase; letter-spacing: 0.06em; }
    .alert-box { background: #fff5f5; border: 1px solid #ffb3b3; border-radius: 6px; padding: 12px 16px; margin: 6px 0; color: #cc0000; font-size: 0.9rem; }
    .alert-box strong { color: #990000; }
    .success-box { background: #f0fff8; border: 1px solid #00c896; border-radius: 6px; padding: 12px 16px; margin: 6px 0; color: #007a5c; font-size: 0.9rem; }
    .info-box { background: #f0f8ff; border: 1px solid #90cce8; border-radius: 6px; padding: 12px 16px; margin: 6px 0; color: #0066aa; font-size: 0.9rem; }
    .stat-card { background: #ffffff; border: 1px solid #dde3ea; border-radius: 8px; padding: 16px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.05); }
    .stat-card .number { color: #00b4d8; font-size: 2rem; font-weight: 700; line-height: 1; }
    .stat-card .label { color: #888; font-size: 0.78rem; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.05em; }
    .step-badge { display: inline-block; background: #00b4d8; color: #ffffff; border-radius: 50%; width: 22px; height: 22px; text-align: center; line-height: 22px; font-weight: 700; font-size: 0.78rem; margin-right: 8px; }
    .modo-muestras { background: linear-gradient(135deg, #fff8e1 0%, #fff3cd 100%); border-left: 5px solid #ffc107; border-radius: 8px; padding: 12px 18px; margin-bottom: 16px; color: #856404; font-weight: 600; font-size: 0.95rem; }
    [data-testid="stToolbar"] { visibility: hidden !important; }
    [data-testid="stDecoration"] { display: none !important; }
    a[href*="github.com"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────
st.markdown("""
<div class="header-box">
    <h1>📋 Generador de Anexo ANMAT</h1>
    <p>Natura · Avon · Operaciones de importación</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# FUNCIONES CORE — OPERACIÓN NORMAL
# ─────────────────────────────────────────

def limpiar_str(s):
    s = str(s).strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.lower()
    s = s.replace(':', ' ').replace('.', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def normalizar_pais(origen_str):
    if pd.isna(origen_str):
        return ''
    s = str(origen_str).strip()
    if ':' in s:
        return s.split(':')[0].strip().lower()
    return s.split(' ')[0].strip().lower()

@st.cache_data
def cargar_anmat(file_bytes):
    buf = BytesIO(file_bytes)
    try:
        df = pd.read_excel(buf, sheet_name='HISTORICO', header=0, engine='pyxlsb')
    except:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name='HISTORICO', header=0)
    df['CM'] = df['CM'].astype(str).str.strip()
    return df

@st.cache_data
def cargar_avon(file_bytes):
    buf = BytesIO(file_bytes)
    return pd.read_excel(buf, header=0)

@st.cache_data
def cargar_fabricantes(file_bytes, suffix='.xlsx'):
    buf = BytesIO(file_bytes)
    if suffix == '.xls':
        df = pd.read_excel(buf, header=1, engine='xlrd')
    else:
        df = pd.read_excel(buf, header=1)
    df.columns = ['material', 'En Historico', 'Corresponde']
    return df

@st.cache_data
def cargar_ncm(file_bytes):
    buf = BytesIO(file_bytes)
    df = pd.read_excel(buf, header=0)
    df['Artículo'] = df['Artículo'].astype(str).str.strip()
    return df

def cargar_pl(file_bytes):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(file_bytes)
        tmp = f.name
    xl = pd.ExcelFile(tmp)
    rows = []
    invoice = None
    for sh in xl.sheet_names:
        df = pd.read_excel(tmp, sheet_name=sh, header=None)
        header_row = None
        for i, row in df.iterrows():
            if 'MATERIAL CODE' in str(row.values):
                header_row = i
                break
        if header_row is None:
            continue
        if not invoice:
            for i, row in df.iterrows():
                for val in row.values:
                    if 'Nº INVOICE:' in str(val) or 'N° INVOICE:' in str(val):
                        idx = list(row.values).index(val)
                        if idx + 1 < len(row.values):
                            invoice = str(row.values[idx + 1]).strip()
                        break
                if invoice:
                    break
        data = df.iloc[header_row + 2:].copy().reset_index(drop=True)
        data.columns = range(len(data.columns))
        data = data[data[1].astype(str).str.match(r'^\d{5,}$')]
        rows.append(data)
    pl = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    return pl, invoice

def cargar_proximas(file_bytes):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(file_bytes)
        tmp = f.name
    df = pd.read_excel(tmp, header=0)
    col_map = {c.strip().lower(): c for c in df.columns}
    if 'material' in col_map:
        df = df.rename(columns={col_map['material']: 'Material'})
    df['Material'] = df['Material'].astype(str).str.strip()
    return df

def buscar_anmat(mat_code, df_anmat):
    found = df_anmat[df_anmat['CM'] == str(mat_code)]
    if len(found) == 0:
        return None
    if len(found) > 1:
        found = found.sort_values('Fecha Admision', ascending=False)
    return found.iloc[0]

def buscar_avon(mat_code, df_avon):
    mat_str = str(mat_code).strip()
    col_cm = 'CM / ZPAC'
    col_fi = 'FI Code Local'
    found = df_avon[df_avon[col_cm].astype(str).str.strip() == mat_str]
    if len(found) == 0:
        found = df_avon[df_avon[col_fi].astype(str).str.strip() == mat_str]
    return found.iloc[0] if len(found) > 0 else None

def buscar_fabricante(origen_str, mat_code, df_fab):
    origen = str(origen_str).strip()
    origen_limpio = limpiar_str(origen)
    match_norm = None
    match_parcial = None
    for _, row in df_fab.iterrows():
        en_hist = str(row['En Historico']).strip()
        if not en_hist or en_hist == 'nan':
            continue
        en_hist_limpio = limpiar_str(en_hist)
        if origen == en_hist:
            return row['Corresponde'], None
        if match_norm is None and origen_limpio == en_hist_limpio:
            match_norm = row['Corresponde']
        if match_parcial is None and en_hist_limpio in origen_limpio:
            match_parcial = row['Corresponde']
    if match_norm:
        return match_norm, None
    if match_parcial:
        return match_parcial, None
    return None, f"Fabricante no encontrado para material {mat_code} (origen: {origen_str})"

def buscar_ncm(mat_code, df_ncm):
    found = df_ncm[df_ncm['Artículo'] == str(mat_code)]
    if len(found) == 0:
        return None, f"NCM no encontrado para material {mat_code}"
    return found.iloc[0]['NCM'], None

def verificar_origen_proximas(origen_anmat, mat_code, df_prox):
    prox_row = df_prox[df_prox['Material'] == str(mat_code)]
    if len(prox_row) == 0:
        return None, f"Material {mat_code} no encontrado en Próximas Importaciones"
    origen_prox = str(prox_row.iloc[0]['Origen'])
    pais_anmat = normalizar_pais(origen_anmat)
    pais_prox = normalizar_pais(origen_prox)
    if pais_anmat != pais_prox:
        return None, f"Origen no coincide para {mat_code}: ANMAT='{origen_anmat}' vs ProxImp='{origen_prox}'"
    return origen_prox, None

SEPARADORES_REGISTRO = [' - ', ' + ', ' / ', ' | ', '\n', '; ', '+', ',']

def separar_registros(registro_str):
    if not registro_str or registro_str == 'nan':
        return [registro_str]
    s = str(registro_str).strip()
    for sep in SEPARADORES_REGISTRO:
        if sep in s:
            partes = [p.strip() for p in s.split(sep) if p.strip()]
            if len(partes) > 1:
                return partes
    return [s]

def buscar_por_registro(nro_registro, df_anmat):
    nro = str(nro_registro).strip()
    found = df_anmat[df_anmat['Registros ANMAT'].astype(str).str.strip() == nro]
    if len(found) == 0:
        return None, "NOT_FOUND"
    if len(found) == 1:
        return [found.iloc[0]], None
    return [found.iloc[i] for i in range(len(found))], "MULTIPLE"

def parsear_fecha_vencimiento(expire_str):
    try:
        if not expire_str or expire_str == 'nan':
            return None
        partes = str(expire_str).strip().split('/')
        if len(partes) == 2:
            mes, anio = int(partes[0]), int(partes[1])
            return datetime(anio, mes, 1)
        return None
    except:
        return None

def verificar_vencimiento(expire_str):
    fecha = parsear_fecha_vencimiento(expire_str)
    if fecha is None:
        return 'ok', None
    hoy = datetime.now()
    limite_90 = hoy + timedelta(days=90)
    if fecha < hoy:
        return 'vencido', f"⚠️ VENCIDO: {expire_str}"
    if fecha <= limite_90:
        return 'proximo', f"⚠️ Vence próximo en 90 días: {expire_str}"
    return 'ok', None

def procesar_pl(pl, df_anmat, df_avon, df_prox, df_fab, df_ncm):
    filas = []
    alertas_excluir = []
    alertas_avon = []
    alertas_generales = []

    for _, pl_row in pl.iterrows():
        mat_code = str(pl_row[1]).strip()
        cantidad = pl_row[2]
        descripcion_pl = str(pl_row[3]).strip() if pd.notna(pl_row[3]) else ''
        lot_product = str(pl_row[5]).strip() if pd.notna(pl_row[5]) else ''
        expire_date = str(pl_row[6]).strip() if pd.notna(pl_row[6]) else ''

        fila = {
            'MATERIAL': mat_code,
            'descripcion_factura': descripcion_pl,
            'Marca y Nombre del producto': '',
            'Variedades': '',
            'Presentación': '',
            'Cantidad': cantidad,
            'N° de inscripcion': '',
            'Lote': lot_product,
            'Fecha de vencimiento': expire_date,
            'Origen': '',
            'Fabricante': '',
            'Posición Arancelaria': '',
            '_alertas': [],
            '_skip': False,
            '_avon': False,
            '_necesita_completar': False,
            '_vencimiento': None,
            '_multi_registro': False,
            '_expanded': False,
        }

        estado_venc, msg_venc = verificar_vencimiento(expire_date)
        if msg_venc:
            fila['_vencimiento'] = estado_venc
            fila['_alertas'].append(msg_venc)
            alertas_generales.append(f"{mat_code} — {msg_venc}")

        anmat_row = buscar_anmat(mat_code, df_anmat)

        if anmat_row is not None:
            nombre = str(anmat_row['NOMBRE']) if pd.notna(anmat_row['NOMBRE']) else ''
            variedad = str(anmat_row['Variedad']) if pd.notna(anmat_row['Variedad']) else ''
            contenido = str(anmat_row['CONTENIDO NETO']) if pd.notna(anmat_row['CONTENIDO NETO']) else ''
            registro = str(anmat_row['Registros ANMAT']) if pd.notna(anmat_row['Registros ANMAT']) else ''
            origen = str(anmat_row['ORIGEN']) if pd.notna(anmat_row['ORIGEN']) else ''

            if 'REFIL' in descripcion_pl.upper():
                nombre = nombre + ' (REPUESTO)'

            fila['Origen'] = normalizar_pais(origen).capitalize() if origen != 'nan' else ''

            _, alerta_origen = verificar_origen_proximas(origen, mat_code, df_prox)
            if alerta_origen:
                fila['_alertas'].append(alerta_origen)
                alertas_generales.append(alerta_origen)

            fab, alerta_fab = buscar_fabricante(origen, mat_code, df_fab)
            if alerta_fab:
                fila['_alertas'].append(alerta_fab)
                alertas_generales.append(alerta_fab)
            else:
                fila['Fabricante'] = fab

            registros = separar_registros(registro)
            if len(registros) <= 1:
                fila['Marca y Nombre del producto'] = nombre
                fila['Variedades'] = variedad if variedad != 'nan' else ''
                fila['Presentación'] = contenido if contenido != 'nan' else ''
                fila['N° de inscripcion'] = registro if registro != 'nan' else ''
            else:
                fila['Marca y Nombre del producto'] = nombre
                fila['Variedades'] = variedad if variedad != 'nan' else ''
                fila['Presentación'] = contenido if contenido != 'nan' else ''
                fila['N° de inscripcion'] = registro if registro != 'nan' else ''
                fila['_multi_registro'] = True
                idx_fila_principal = len(filas)
                filas.append(fila)
                for nro in registros:
                    anmat_rows, status = buscar_por_registro(nro, df_anmat)
                    if status == "NOT_FOUND":
                        msg = "No encontrado: " + nro
                        fila_exp = {
                            'MATERIAL': '', 'descripcion_factura': '',
                            'Marca y Nombre del producto': '', 'Variedades': '',
                            'Presentación': '', 'Cantidad': '',
                            'N° de inscripcion': nro, 'Lote': '',
                            'Fecha de vencimiento': '', 'Origen': '',
                            'Fabricante': '', 'Posición Arancelaria': '',
                            '_alertas': [msg], '_skip': False,
                            '_avon': False, '_necesita_completar': False,
                            '_expanded': True, '_multi_opciones': False,
                            '_nro_registro': nro,
                        }
                        alertas_generales.append(msg)
                        filas.append(fila_exp)
                    else:
                        es_multiple = status == "MULTIPLE"
                        for multi_i, anmat_nro in enumerate(anmat_rows):
                            n = str(anmat_nro['NOMBRE']) if pd.notna(anmat_nro['NOMBRE']) else ''
                            v = str(anmat_nro['Variedad']) if pd.notna(anmat_nro['Variedad']) else ''
                            c = str(anmat_nro['CONTENIDO NETO']) if pd.notna(anmat_nro['CONTENIDO NETO']) else ''
                            if 'REFIL' in descripcion_pl.upper():
                                n = n + ' (REPUESTO)'
                            fila_exp = {
                                'MATERIAL': '', 'descripcion_factura': '',
                                'Marca y Nombre del producto': n,
                                'Variedades': v if v != 'nan' else '',
                                'Presentación': c if c != 'nan' else '',
                                'Cantidad': '', 'N° de inscripcion': nro,
                                'Lote': '', 'Fecha de vencimiento': '',
                                'Origen': '', 'Fabricante': '',
                                'Posición Arancelaria': '',
                                '_alertas': [], '_skip': es_multiple,
                                '_avon': False, '_necesita_completar': False,
                                '_expanded': True, '_multi_opciones': es_multiple,
                                '_nro_registro': nro,
                                '_multi_idx': multi_i,
                            }
                            filas.append(fila_exp)
                ncm, alerta_ncm = buscar_ncm(mat_code, df_ncm)
                if alerta_ncm:
                    alertas_generales.append(alerta_ncm)
                else:
                    filas[idx_fila_principal]['Posición Arancelaria'] = ncm
                continue

        else:
            avon_row = buscar_avon(mat_code, df_avon)
            if avon_row is not None:
                fila['_avon'] = True
                cm_zpac = str(avon_row.get('CM / ZPAC', '')).strip()
                fi_code = str(avon_row.get('FI Code Local', '')).strip()
                fila['MATERIAL'] = cm_zpac if cm_zpac and cm_zpac != 'nan' else fi_code

                nombre_avon = str(avon_row.get('NOMBRE DE REGISTRO DE PRODUCTO', ''))
                contenido_avon = str(avon_row.get('CONTENIDO LEGAL', ''))
                registro_avon = str(avon_row.get('Reg. SP   (Trámite#)\nARGENTINA NATURA', ''))

                if 'REFIL' in descripcion_pl.upper():
                    nombre_avon = nombre_avon + ' (REPUESTO)'

                fila['Marca y Nombre del producto'] = nombre_avon if nombre_avon != 'nan' else ''
                fila['Presentación'] = contenido_avon if contenido_avon != 'nan' else ''
                fila['N° de inscripcion'] = registro_avon if registro_avon != 'nan' else ''
                fila['Variedades'] = ''
                fila['Origen'] = ''
                fila['_necesita_completar'] = True

                fab, alerta_fab = buscar_fabricante('', mat_code, df_fab)
                fila['Fabricante'] = fab if not alerta_fab else ''

                alertas_avon.append({
                    'material': mat_code,
                    'descripcion': descripcion_pl,
                    'fila_idx': len(filas)
                })
            else:
                fila['_skip'] = True
                alertas_excluir.append({
                    'material': mat_code,
                    'descripcion': descripcion_pl,
                    'fila_idx': len(filas)
                })

        ncm, alerta_ncm = buscar_ncm(mat_code, df_ncm)
        if alerta_ncm:
            fila['_alertas'].append(alerta_ncm)
            alertas_generales.append(alerta_ncm)
            fila['Posición Arancelaria'] = ''
        else:
            fila['Posición Arancelaria'] = ncm

        filas.append(fila)

    return filas, alertas_excluir, alertas_avon, alertas_generales

def separar_anexos(filas):
    principal, difusor, kit3x1, alertas_sep = [], [], [], []
    for fila in filas:
        if fila['_skip']:
            continue
        desc = fila['descripcion_factura'].upper()
        es_difusor = 'DIFUSOR' in desc
        es_3x1 = bool(re.search(r'3\s*[Xx]\s*1(?![0-9])', desc))
        if es_difusor and es_3x1:
            alertas_sep.append(f"Material {fila['MATERIAL']} tiene DIFUSOR y 3X1 — verificar.")
            principal.append(fila)
        elif es_difusor:
            difusor.append(fila)
        elif es_3x1:
            kit3x1.append(fila)
        else:
            principal.append(fila)
    return principal, difusor, kit3x1, alertas_sep

# ─────────────────────────────────────────
# FUNCIONES MUESTRAS NATURA
# ─────────────────────────────────────────

FABRICANTE_MUESTRAS = 'INDUSTRIA E COMERCIO DE COSMÉTICOS NATURA LTDA'
ORIGEN_MUESTRAS = 'Brasil'

def parsear_msg(file_bytes):
    """
    Lee un archivo .msg y extrae la tabla Código / NCM / ANMAT
    del primer bloque del cuerpo (antes del primer separador de reply).
    Retorna: (lista de dicts {'codigo', 'ncm', 'anmat'}, error_str o None)
    """
    try:
        import extract_msg
        with tempfile.NamedTemporaryFile(suffix='.msg', delete=False) as f:
            f.write(file_bytes)
            tmp = f.name
        msg = extract_msg.Message(tmp)
        body = msg.body or ''
    except Exception as e:
        return None, f"No se pudo leer el archivo .msg: {e}"

    # Solo primer bloque (antes del primer separador de cadena de reply)
    primer_bloque = re.split(r'_{3,}', body)[0]
    lineas = [l.strip() for l in primer_bloque.replace('\r\n', '\n').split('\n')]

    # Buscar inicio de tabla
    inicio = None
    for i, l in enumerate(lineas):
        if 'código del artículo' in l.lower() or 'codigo del articulo' in l.lower():
            inicio = i + 1
            break

    if inicio is None:
        return None, "No se encontró la tabla Código / NCM / ANMAT en el mail."

    # Filtrar líneas vacías y nombres de columnas repetidos
    cols_ignorar = {'ncm', 'anmat', 'código del artículo', 'codigo del articulo', ''}
    tokens = [l for l in lineas[inicio:] if l.lower() not in cols_ignorar]

    # Agrupar de a 3: codigo, ncm, si/no
    items = []
    i = 0
    while i + 2 < len(tokens):
        codigo = tokens[i].strip()
        ncm    = tokens[i+1].strip()
        anmat_val = tokens[i+2].strip().lower()
        if anmat_val in ('si', 'sí', 'no') and re.match(r'^\d{4,}', ncm):
            items.append({
                'codigo': codigo,
                'ncm':    ncm,
                'anmat':  anmat_val in ('si', 'sí')
            })
            i += 3
        else:
            i += 1  # desplazar si no matchea el patrón

    if not items:
        return None, "La tabla del mail no tiene el formato esperado (Código / NCM / Sí o No)."

    return items, None


def cargar_pl_muestras(file_bytes):
    """
    Carga el Packing List para muestras.
    Detecta TODAS las columnas de cantidad presentes (KG, gramos, unidades).
    Pueden coexistir las tres en el mismo archivo.
    Retorna: (lista de dicts por ítem, invoice str)
    """
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(file_bytes)
        tmp = f.name

    xl = pd.ExcelFile(tmp)
    invoice = None
    items = []

    for sh in xl.sheet_names:
        df_raw = pd.read_excel(tmp, sheet_name=sh, header=None)

        # Buscar invoice
        if not invoice:
            for i, row in df_raw.iterrows():
                for val in row.values:
                    if 'Nº INVOICE:' in str(val) or 'N° INVOICE:' in str(val):
                        idx_v = list(row.values).index(val)
                        if idx_v + 1 < len(row.values):
                            invoice = str(row.values[idx_v + 1]).strip()
                        break
                if invoice:
                    break

        # Buscar fila de headers con MATERIAL CODE
        header_row_idx = None
        for i, row in df_raw.iterrows():
            if any('MATERIAL CODE' in str(v).upper() for v in row.values if v):
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        headers = [str(v).strip().upper() if v else '' for v in df_raw.iloc[header_row_idx].values]

        # Mapear columnas — pueden coexistir KG, gramos y unidades
        col_material = None
        col_qty_kg   = None
        col_qty_g    = None
        col_qty_un   = None
        col_desc     = None
        col_lot      = None
        col_expire   = None

        for idx, h in enumerate(headers):
            if 'MATERIAL CODE' in h:
                col_material = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and ('KG' in h or 'KILO' in h):
                if col_qty_kg is None:
                    col_qty_kg = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and ('GRAM' in h) and 'KG' not in h:
                if col_qty_g is None:
                    col_qty_g = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and ('UNIT' in h or 'UNID' in h):
                if col_qty_un is None:
                    col_qty_un = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and col_qty_kg is None and col_qty_g is None and col_qty_un is None:
                # Primera columna de cantidad sin especificador → tratarla como KG
                col_qty_kg = idx
            elif 'DESCRIPTION' in h or 'DESCRIPCI' in h:
                if col_desc is None:
                    col_desc = idx
            elif 'LOT' in h or 'LOTE' in h:
                if col_lot is None:
                    col_lot = idx
            elif 'EXPIRE' in h or 'VENC' in h:
                if col_expire is None:
                    col_expire = idx

        if col_material is None:
            continue

        # Leer filas de datos (saltear fila vacía inmediata debajo del header)
        data_start = header_row_idx + 2
        for i in range(data_start, len(df_raw)):
            row = df_raw.iloc[i]
            mat = str(row.iloc[col_material]).strip() if col_material is not None else ''
            if not mat or mat == 'nan' or not re.search(r'\d', mat):
                continue
            # Detener al llegar a filas de totales / observaciones
            if any(kw in mat.upper() for kw in ('VOLUME', 'OBSERVAC', 'TOTAL')):
                continue

            desc = str(row.iloc[col_desc]).strip() if col_desc is not None and pd.notna(row.iloc[col_desc]) else ''
            lot  = str(row.iloc[col_lot]).strip()  if col_lot  is not None and pd.notna(row.iloc[col_lot])  else ''

            # Fecha de vencimiento
            expire_raw = row.iloc[col_expire] if col_expire is not None else None
            if isinstance(expire_raw, datetime):
                expire_str = expire_raw.strftime('%d/%m/%Y')
            elif expire_raw is not None and str(expire_raw) != 'nan':
                expire_str = str(expire_raw).strip()
            else:
                expire_str = ''

            # Cantidades — capturar todos los tipos presentes
            cantidades = {}
            if col_qty_kg is not None:
                v = row.iloc[col_qty_kg]
                if pd.notna(v) and str(v).strip() not in ('nan', ''):
                    try:
                        cantidades['kg'] = float(v)
                    except:
                        pass
            if col_qty_g is not None:
                v = row.iloc[col_qty_g]
                if pd.notna(v) and str(v).strip() not in ('nan', ''):
                    try:
                        cantidades['g'] = float(v)
                    except:
                        pass
            if col_qty_un is not None:
                v = row.iloc[col_qty_un]
                if pd.notna(v) and str(v).strip() not in ('nan', ''):
                    try:
                        cantidades['un'] = float(v)
                    except:
                        pass

            items.append({
                'material':    mat,
                'descripcion': desc,
                'lot':         lot,
                'expire':      expire_str,
                'cantidades':  cantidades,
            })

    return items, invoice


def _resolver_presentacion_cantidad(cantidades):
    """
    Dado el dict de cantidades detectadas, retorna (presentacion, cantidad_valor).

    Reglas:
    - 1 tipo  → presentacion = 'kilos'/'gramos'/'unidades', cantidad = valor numérico
    - N tipos → presentacion = 'kilos / gramos / unidades' (los que correspondan),
                cantidad     = 'X kg / Y g / Z un' (texto con la unidad en cada valor)
    """
    nombre_map  = {'kg': 'kilos', 'g': 'gramos', 'un': 'unidades'}
    sufijo_map  = {'kg': 'kg',    'g': 'g',       'un': 'un'}
    tipos = [t for t in ('kg', 'g', 'un') if t in cantidades]

    if len(tipos) == 0:
        return '', ''
    if len(tipos) == 1:
        t = tipos[0]
        v = cantidades[t]
        # Mostrar como entero si no tiene decimales
        val = int(v) if v == int(v) else v
        return nombre_map[t], val
    # Varios tipos
    presentacion = ' / '.join(nombre_map[t] for t in tipos)
    cantidad_str = ' / '.join(f"{int(cantidades[t]) if cantidades[t] == int(cantidades[t]) else cantidades[t]} {sufijo_map[t]}" for t in tipos)
    return presentacion, cantidad_str


def procesar_muestras(items_pl, items_mail):
    """
    Cruza PL con la tabla del mail.
    Solo procesa los ítems donde ANMAT = Sí.
    NCM viene del mail.
    Retorna: (filas_anexo, codigos_si_no_encontrados_en_pl)
    """
    mail_dict   = {it['codigo'].strip(): it for it in items_mail}
    codigos_si  = {cod for cod, it in mail_dict.items() if it['anmat']}
    codigos_pl  = {it['material'] for it in items_pl}

    filas = []
    for item in items_pl:
        mat = item['material']
        if mat not in codigos_si:
            continue  # ANMAT = No o no está en el mail → ignorar

        mail_item = mail_dict[mat]
        presentacion, cantidad = _resolver_presentacion_cantidad(item['cantidades'])

        filas.append({
            'MATERIAL':                     mat,
            'descripcion_factura':          item['descripcion'],
            'Marca y Nombre del producto':  item['descripcion'],
            'Variedades':                   'N/C',
            'Presentación':                 presentacion,
            'Cantidad':                     cantidad,
            'N° de inscripcion':            'N/C',
            'Lote':                         item['lot'],
            'Fecha de vencimiento':         item['expire'],
            'Origen':                       ORIGEN_MUESTRAS,
            'Fabricante':                   FABRICANTE_MUESTRAS,
            'Posición Arancelaria':         mail_item['ncm'],
            # Internos
            '_alertas':           [],
            '_skip':              False,
            '_avon':              False,
            '_necesita_completar': False,
            '_vencimiento':       None,
            '_multi_registro':    False,
            '_expanded':          False,
        })

    # Códigos con ANMAT=Sí que no aparecieron en el PL
    no_en_pl = [cod for cod in codigos_si if cod not in codigos_pl]

    return filas, no_en_pl


# ─────────────────────────────────────────
# FUNCIONES DE SALIDA (compartidas)
# ─────────────────────────────────────────

COLUMNAS_SALIDA = ['MATERIAL', 'descripcion_factura', 'Marca y Nombre del producto',
                   'Variedades', 'Presentación', 'Cantidad', 'N° de inscripcion',
                   'Lote', 'Fecha de vencimiento', 'Origen', 'Fabricante', 'Posición Arancelaria']
COLUMNAS_SIN_PRIMERAS = COLUMNAS_SALIDA[2:]

ANCHOS = {'MATERIAL': 12, 'descripcion_factura': 35, 'Marca y Nombre del producto': 45,
          'Variedades': 20, 'Presentación': 14, 'Cantidad': 14, 'N° de inscripcion': 22,
          'Lote': 14, 'Fecha de vencimiento': 18, 'Origen': 14, 'Fabricante': 45,
          'Posición Arancelaria': 20}

def escribir_excel_bytes(filas, incluir_primeras_cols=True):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Anexo de Productos'
    columnas = COLUMNAS_SALIDA if incluir_primeras_cols else COLUMNAS_SIN_PRIMERAS

    ws.merge_cells(f'A1:{get_column_letter(len(columnas))}1')
    titulo = ws['A1']
    titulo.value = 'ANEXO DE PRODUCTOS'
    titulo.font = Font(name='Arial', bold=True, size=11)
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    titulo.fill = PatternFill('solid', start_color='D9D9D9')

    header_fill = PatternFill('solid', start_color='70AD47')
    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    alerta_fill = PatternFill('solid', start_color='FFEB9C')
    for row_idx, fila in enumerate(filas, 3):
        tiene_alerta = len(fila.get('_alertas', [])) > 0 or fila.get('_necesita_completar', False)
        for col_idx, col_name in enumerate(columnas, 1):
            val = fila.get(col_name, '')
            # MATERIAL como número entero
            if col_name == 'MATERIAL' and val != '':
                try:
                    val = int(float(str(val)))
                except:
                    pass
            # Cantidad: número si es puro, string si tiene unidades mezcladas
            if col_name == 'Cantidad' and val != '':
                if not isinstance(val, str):
                    try:
                        v = float(val)
                        val = int(v) if v == int(v) else v
                    except:
                        pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if tiene_alerta:
                cell.fill = alerta_fill

    for col_idx, col_name in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ANCHOS.get(col_name, 15)

    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'A3'

    if not incluir_primeras_cols:
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def excel_a_pdf_bytes(excel_bytes, nombre_base):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb.active

        style_normal = ParagraphStyle('normal', fontSize=6.5, leading=8, alignment=TA_LEFT)
        style_header = ParagraphStyle('header', fontSize=7, leading=9, alignment=TA_CENTER,
                                       textColor=colors.white, fontName='Helvetica-Bold')
        style_title  = ParagraphStyle('title',  fontSize=8, leading=10, alignment=TA_CENTER,
                                       fontName='Helvetica-Bold')

        ancho_total = landscape(A4)[0] - 1.4*cm
        pesos = [4.2, 1.8, 1.1, 0.8, 2.0, 0.9, 1.4, 1.1, 3.2, 1.8]
        total_pesos = sum(pesos)
        col_widths = [ancho_total * p / total_pesos for p in pesos]

        def safe_para(val, style):
            try:
                txt = str(val) if val is not None else ''
                txt = txt.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                return Paragraph(txt, style)
            except:
                return Paragraph('', style)

        data = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                data.append([safe_para(row[0], style_title)] + ['' for _ in range(len(pesos)-1)])
            elif row_idx == 1:
                cells = list(row)
                while len(cells) < len(pesos): cells.append('')
                cells = cells[:len(pesos)]
                data.append([safe_para(c, style_header) for c in cells])
            else:
                cells = list(row)
                while len(cells) < len(pesos): cells.append('')
                cells = cells[:len(pesos)]
                data.append([safe_para(c, style_normal) for c in cells])

        if not data:
            return None

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=0.7*cm, rightMargin=0.7*cm,
                                topMargin=0.8*cm, bottomMargin=0.8*cm)
        table = Table(data, colWidths=col_widths, repeatRows=2)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D9D9D9')),
            ('SPAN',       (0,0), (-1,0)),
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#70AD47')),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0,2), (-1,-1), [colors.white, colors.HexColor('#F7F7F7')]),
            ('LEFTPADDING',   (0,0), (-1,-1), 3),
            ('RIGHTPADDING',  (0,0), (-1,-1), 3),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        doc.build([table])
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        print(f'PDF error: {e}')
        return None

def generar_zip(grupos, invoice):
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nombre, filas in grupos:
            if not filas:
                continue
            nombre_base = f'ANEXO_{nombre}_{invoice}'
            xls_completo = escribir_excel_bytes(filas, incluir_primeras_cols=True)
            zf.writestr(f'{nombre_base}.xlsx', xls_completo)
            xls_sin = escribir_excel_bytes(filas, incluir_primeras_cols=False)
            zf.writestr(f'{nombre_base}_SIN_MAT.xlsx', xls_sin)
            pdf = excel_a_pdf_bytes(xls_sin, f'{nombre_base}_SIN_MAT')
            if pdf:
                zf.writestr(f'{nombre_base}_SIN_MAT.pdf', pdf)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────
defaults = {
    # operación normal
    'filas_procesadas':        None,
    'alertas_excluir':         [],
    'alertas_avon':            [],
    'alertas_generales':       [],
    'invoice':                 None,
    'excluidos':               set(),
    'datos_avon_completados':  {},
    # muestras
    'filas_muestras':          None,
    'invoice_muestras':        None,
    'alertas_muestras':        [],
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────
# PASO 0: SELECCIÓN DE MODO
# ─────────────────────────────────────────
st.markdown('<div class="card"><h3><span class="step-badge">0</span>Tipo de operación</h3>', unsafe_allow_html=True)
modo = st.radio(
    "¿Qué tipo de operación es?",
    options=["Operación normal", "Muestras Natura"],
    horizontal=True,
    key='modo_radio'
)
st.markdown('</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════
# RAMA A: MUESTRAS NATURA
# ═══════════════════════════════════════════════════════════
if modo == "Muestras Natura":

    st.markdown('<div class="modo-muestras">🧪 Modo Muestras Natura — se generará el Anexo solo para los ítems con ANMAT = Sí del mail de clasificación. NCM, origen y fabricante se toman automáticamente.</div>', unsafe_allow_html=True)

    # ── Paso 1: Archivos ──
    st.markdown('<div class="card"><h3><span class="step-badge">1</span>Archivos de la operación</h3>', unsafe_allow_html=True)
    st.markdown("**📌 Número de referencia de la operación**")
    nro_ref_m = st.text_input("", placeholder="ej: MN014-26", label_visibility="collapsed", key='nro_ref_muestras')

    col1, col2 = st.columns(2)
    with col1:
        f_pl_m = st.file_uploader("📦 Packing List / Invoice (.xlsx)", type=['xlsx'], key='pl_muestras')
    with col2:
        f_msg  = st.file_uploader("📧 Mail de clasificación ANMAT (.msg)", type=['msg'], key='msg_muestras')
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Paso 2: Procesar ──
    if f_pl_m and f_msg:
        st.markdown('<div class="card"><h3><span class="step-badge">2</span>Procesar</h3>', unsafe_allow_html=True)
        if st.button("⚙️ Analizar y procesar muestras", key='btn_procesar_muestras'):
            with st.spinner('Procesando...'):
                try:
                    items_mail, err_mail = parsear_msg(f_msg.read())
                    if err_mail:
                        st.error(f"Error al leer el mail: {err_mail}")
                        st.stop()

                    items_pl, invoice_m = cargar_pl_muestras(f_pl_m.read())
                    if not items_pl:
                        st.error("No se encontraron ítems en el Packing List.")
                        st.stop()

                    filas_m, no_en_pl = procesar_muestras(items_pl, items_mail)

                    alertas_m = []
                    for cod in no_en_pl:
                        alertas_m.append(f"⚠️ Código {cod} tiene ANMAT=Sí en el mail pero no se encontró en el Packing List.")

                    st.session_state.filas_muestras    = filas_m
                    st.session_state.invoice_muestras  = invoice_m
                    st.session_state.alertas_muestras  = alertas_m

                except Exception as e:
                    import traceback
                    st.error(f"Error al procesar: {e}")
                    st.text(traceback.format_exc())
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Paso 3: Resultados ──
    if st.session_state.filas_muestras is not None:
        filas_m   = st.session_state.filas_muestras
        invoice_m = st.session_state.invoice_muestras

        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f'<div class="stat-card"><div class="number">{len(filas_m)}</div><div class="label">Ítems con ANMAT = Sí</div></div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="stat-card"><div class="number" style="color:#00c896">{len(filas_m)}</div><div class="label">Líneas en el Anexo</div></div>', unsafe_allow_html=True)

        st.markdown('<br>', unsafe_allow_html=True)

        if st.session_state.alertas_muestras:
            st.markdown('<div class="card"><h3>⚠️ Alertas</h3>', unsafe_allow_html=True)
            for a in st.session_state.alertas_muestras:
                st.markdown(f'<div class="alert-box">{a}</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        with st.expander("👁️ Vista previa del Anexo de Muestras"):
            cols_preview = ['MATERIAL', 'Marca y Nombre del producto', 'Presentación',
                            'Cantidad', 'Lote', 'Fecha de vencimiento', 'Posición Arancelaria']
            st.dataframe(pd.DataFrame([{c: f.get(c,'') for c in cols_preview} for f in filas_m]),
                         use_container_width=True)

        # ── Paso 4: Generar ──
        st.markdown('<div class="card"><h3><span class="step-badge">4</span>Generar Anexo de Muestras</h3>', unsafe_allow_html=True)
        if st.button("📄 Generar Anexo de Muestras", key='btn_generar_muestras'):
            with st.spinner('Generando archivos...'):
                ref = nro_ref_m.strip() if nro_ref_m.strip() else (invoice_m or 'MUESTRAS')
                zip_bytes = generar_zip([('MUESTRAS', filas_m)], ref)
                st.markdown('<div class="success-box">✅ Anexo de Muestras generado correctamente</div>', unsafe_allow_html=True)
                st.markdown(f"**MUESTRAS**: {len(filas_m)} ítems")
                st.download_button(
                    label="⬇️ Descargar Anexo de Muestras (ZIP)",
                    data=zip_bytes,
                    file_name=f"ANEXO_MUESTRAS_{ref}.zip",
                    mime="application/zip",
                    key='dl_zip_muestras'
                )
        st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════
# RAMA B: OPERACIÓN NORMAL
# ═══════════════════════════════════════════════════════════
else:

    # ── Paso 1: Archivos ──
    st.markdown('<div class="card"><h3><span class="step-badge">1</span>Archivos de la operación</h3>', unsafe_allow_html=True)
    st.markdown("**📌 Número de referencia de la operación**")
    nro_referencia = st.text_input("", placeholder="ej: 4550595912", label_visibility="collapsed")

    col1, col2 = st.columns(2)
    with col1:
        f_pl    = st.file_uploader("📦 Packing List",                  type=['xlsx'],        key='pl')
        f_prox  = st.file_uploader("📅 Próximas Importaciones",        type=['xlsx'],        key='prox')
        f_anmat = st.file_uploader("🏥 Registro ANMAT Histórico",      type=['xlsb','xlsx'], key='anmat')
    with col2:
        f_avon  = st.file_uploader("🌸 Registros Avon",               type=['xlsx'],        key='avon')
        f_fab   = st.file_uploader("🏭 Fabricantes",                   type=['xls','xlsx'],  key='fab')
        f_ncm   = st.file_uploader("📊 Catálogo NCM",                  type=['xlsx'],        key='ncm')
    st.markdown('</div>', unsafe_allow_html=True)

    archivos_ok = all([f_pl, f_prox, f_anmat, f_avon, f_fab, f_ncm])

    # ── Paso 2: Procesar ──
    if archivos_ok:
        st.markdown('<div class="card"><h3><span class="step-badge">2</span>Procesar operación</h3>', unsafe_allow_html=True)
        if st.button("⚙️ Analizar y procesar", key='btn_procesar'):
            with st.spinner('Procesando...'):
                try:
                    suffix_fab = '.xls' if f_fab.name.endswith('.xls') else '.xlsx'
                    pl, invoice        = cargar_pl(f_pl.read())
                    df_prox            = cargar_proximas(f_prox.read())
                    df_anmat           = cargar_anmat(f_anmat.read())
                    df_avon            = cargar_avon(f_avon.read())
                    df_fab             = cargar_fabricantes(f_fab.read(), suffix=suffix_fab)
                    df_ncm             = cargar_ncm(f_ncm.read())

                    filas, alertas_excluir, alertas_avon, alertas_generales = procesar_pl(
                        pl, df_anmat, df_avon, df_prox, df_fab, df_ncm
                    )

                    st.session_state.filas_procesadas       = filas
                    st.session_state.alertas_excluir        = alertas_excluir
                    st.session_state.alertas_avon           = alertas_avon
                    st.session_state.alertas_generales      = alertas_generales
                    st.session_state.invoice                = invoice
                    st.session_state.excluidos              = set()
                    st.session_state.datos_avon_completados = {}

                except Exception as e:
                    st.error(f"Error al procesar: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Paso 3: Alertas y resolución ──
    if st.session_state.filas_procesadas is not None:
        filas   = st.session_state.filas_procesadas
        invoice = st.session_state.invoice

        total    = len(filas)
        skip     = len(st.session_state.alertas_excluir)
        avon     = len(st.session_state.alertas_avon)
        generales= len(st.session_state.alertas_generales)

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f'<div class="stat-card"><div class="number">{total}</div><div class="label">Ítems PL</div></div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="stat-card"><div class="number" style="color:#00c896">{total - skip}</div><div class="label">A procesar</div></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="stat-card"><div class="number" style="color:#ffd166">{avon}</div><div class="label">Avon / completar</div></div>', unsafe_allow_html=True)
        with col4:
            st.markdown(f'<div class="stat-card"><div class="number" style="color:#ff6b6b">{skip}</div><div class="label">No encontrados</div></div>', unsafe_allow_html=True)

        st.markdown('<br>', unsafe_allow_html=True)

        # Registros con múltiples coincidencias
        filas_multi = [f for f in st.session_state.filas_procesadas if f.get('_multi_opciones')]
        if filas_multi:
            grupos_multi = {}
            for f in filas_multi:
                nro = f.get('_nro_registro', '')
                if nro not in grupos_multi:
                    grupos_multi[nro] = []
                grupos_multi[nro].append(f)
            if 'incluidos_multi' not in st.session_state:
                st.session_state.incluidos_multi = set()
            st.markdown('<div class="card"><h3>🔀 Registros con múltiples coincidencias</h3>', unsafe_allow_html=True)
            for nro, opciones in grupos_multi.items():
                st.markdown('<div class="alert-box"><strong>Registro ' + nro + '</strong> — encontrado ' + str(len(opciones)) + ' veces. Seleccioná cuál/es incluir:</div>', unsafe_allow_html=True)
                for i, op in enumerate(opciones):
                    key_op  = 'multi_' + nro + '_' + str(i)
                    incluido = key_op in st.session_state.incluidos_multi
                    col1, col2 = st.columns([5, 1])
                    with col1:
                        st.markdown('<div style="background:#f8f9fa;border:1px solid #dde3ea;border-radius:6px;padding:10px;margin:4px 0;font-size:0.85rem;"><strong>' +
                                    op.get('Marca y Nombre del producto','') + '</strong> | Variedad: ' +
                                    op.get('Variedades','—') + ' | Presentación: ' + op.get('Presentación','—') + '</div>', unsafe_allow_html=True)
                    with col2:
                        if st.button("✅ Incluida" if incluido else "Incluir", key='btn_' + key_op):
                            if incluido:
                                st.session_state.incluidos_multi.discard(key_op)
                            else:
                                st.session_state.incluidos_multi.add(key_op)
                            st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        # Vencimientos
        alertas_venc = [f for f in st.session_state.filas_procesadas if f.get('_vencimiento') in ('vencido','proximo') and not f['_skip']]
        if alertas_venc:
            st.markdown('<div class="card"><h3>⏰ Alertas de vencimiento</h3>', unsafe_allow_html=True)
            for fila in alertas_venc:
                tipo    = fila.get('_vencimiento')
                color   = '#cc0000' if tipo == 'vencido' else '#cc7700'
                icono   = '🔴' if tipo == 'vencido' else '🟡'
                msg     = [a for a in fila['_alertas'] if 'venc' in a.lower()]
                msg_str = msg[0] if msg else ''
                key_excl= f'venc_{fila["MATERIAL"]}_{fila["Lote"]}'
                excluido= key_excl in st.session_state.excluidos
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.markdown(f'<div class="alert-box" style="border-color:{color};color:{color};">{icono} <strong>{fila["MATERIAL"]}</strong> — Lote {fila["Lote"]} — {msg_str}</div>', unsafe_allow_html=True)
                with col2:
                    if st.button("✅ Excluido" if excluido else "Excluir", key=f'btn_venc_{key_excl}'):
                        st.session_state.excluidos.add(key_excl)
                        st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        # No encontrados
        if st.session_state.alertas_excluir:
            st.markdown('<div class="card"><h3>⚠️ No encontrados en ANMAT ni Avon — ¿excluir del anexo?</h3>', unsafe_allow_html=True)
            for idx_excl, item in enumerate(st.session_state.alertas_excluir):
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.markdown(f'<div class="alert-box"><strong>{item["material"]}</strong> — {item["descripcion"]}</div>', unsafe_allow_html=True)
                with col2:
                    excluido = item['material'] in st.session_state.excluidos
                    if st.button("✅ Excluido" if excluido else "Excluir del Anexo", key=f'excl_{item["material"]}_{idx_excl}'):
                        st.session_state.excluidos.add(item['material'])
                        st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        # Avon
        if st.session_state.alertas_avon:
            st.markdown('<div class="card"><h3>🌸 Ítems Avon — completar Fabricante y Origen</h3>', unsafe_allow_html=True)
            for item in st.session_state.alertas_avon:
                mat  = item['material']
                st.markdown(f'<div class="alert-box"><strong>{mat}</strong> — {item["descripcion"]}</div>', unsafe_allow_html=True)
                prev = st.session_state.datos_avon_completados.get(mat, {})
                c1, c2, c3 = st.columns(3)
                with c1: fab_val = st.text_input("Fabricante", value=prev.get('fabricante',''), key=f'fab_{mat}')
                with c2: orig_val= st.text_input("Origen",     value=prev.get('origen',''),     key=f'orig_{mat}')
                with c3: var_val = st.text_input("Variedad",   value=prev.get('variedad',''),   key=f'var_{mat}')
                if fab_val or orig_val or var_val:
                    st.session_state.datos_avon_completados[mat] = {'fabricante': fab_val, 'origen': orig_val, 'variedad': var_val}
            st.markdown('</div>', unsafe_allow_html=True)

        # Alertas generales
        if st.session_state.alertas_generales:
            with st.expander(f"⚠️ {len(st.session_state.alertas_generales)} alertas adicionales"):
                for a in st.session_state.alertas_generales:
                    st.markdown(f'<div class="alert-box">{a}</div>', unsafe_allow_html=True)

        # ── Paso 4: Generar ──
        st.markdown('<div class="card"><h3><span class="step-badge">4</span>Generar Anexo</h3>', unsafe_allow_html=True)
        if st.button("📄 Generar Anexo completo", key='btn_generar'):
            with st.spinner('Generando archivos...'):
                incluidos_multi = st.session_state.get('incluidos_multi', set())
                filas_final = []
                for fila in filas:
                    mat      = fila['MATERIAL']
                    key_venc = mat + '_' + fila.get('Lote', '')
                    if mat in st.session_state.excluidos:
                        continue
                    if 'venc_' + key_venc in st.session_state.excluidos:
                        continue
                    if fila.get('_multi_opciones'):
                        nro    = fila.get('_nro_registro', '')
                        idx    = fila.get('_multi_idx', -1)
                        key_op = 'multi_' + nro + '_' + str(idx)
                        if key_op not in incluidos_multi:
                            continue
                        f = fila.copy(); f['_skip'] = False
                    elif fila.get('_skip'):
                        continue
                    else:
                        f = fila.copy()
                    if fila.get('_avon') and mat in st.session_state.datos_avon_completados:
                        datos = st.session_state.datos_avon_completados[mat]
                        if datos.get('fabricante'): f['Fabricante'] = datos['fabricante']
                        if datos.get('origen'):     f['Origen']     = datos['origen']
                        if datos.get('variedad'):   f['Variedades'] = datos['variedad']
                    filas_final.append(f)

                principal, difusor, kit3x1, _ = separar_anexos(filas_final)
                grupos    = [('PRINCIPAL', principal), ('DIFUSOR', difusor), ('3x1', kit3x1)]
                ref       = nro_referencia.strip() if nro_referencia.strip() else invoice
                zip_bytes = generar_zip(grupos, ref)

                st.markdown('<div class="success-box">✅ Anexo generado correctamente</div>', unsafe_allow_html=True)
                resumen = [f"**{n}**: {len(fg)} ítems" for n, fg in grupos if fg]
                st.markdown(' · '.join(resumen))

                st.download_button(
                    label="⬇️ Descargar todos los archivos (ZIP)",
                    data=zip_bytes,
                    file_name=f"ANEXO_{ref}.zip",
                    mime="application/zip",
                    key='dl_zip'
                )
        st.markdown('</div>', unsafe_allow_html=True)
